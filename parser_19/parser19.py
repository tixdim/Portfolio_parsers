import requests
from lxml import etree
import lxml.html
import openpyxl


def parse(url):
    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    words = tree.xpath("/html/body/div[4]/div[1]/div[1]/div[1]/div[2]/ul/li/a/text()")
    return (words)


def main():
    url = "https://www.allscrabblewords.com/{number}-letter-words/"
    i = 1
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    while i <= 12:
        if i == 1:
            letters = []
            for h in range(97, 123):
                letters.append(chr(h))

            wb.create_sheet(title=f"Page {i}")
            sheet = wb[f"Page {i}"]
            sheet.column_dimensions['A'].width = i + 5
            for word in letters:
                cell = sheet.cell(row=letters.index(word) + 1, column=1)
                cell.value = word
            i += 1
        else:
            words = parse(url.format(number=i))
            wb.create_sheet(title=f"Page {i}")
            sheet = wb[f"Page {i}"]
            sheet.column_dimensions['A'].width = i + 5
            for word in words:
                cell = sheet.cell(row=words.index(word) + 1, column=1)
                cell.value = word
            i += 1
    wb.save("pars_words.xlsx")


if __name__ == "__main__":
    main()